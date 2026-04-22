import base64
import binascii
import json
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

from models import (
    SimulationSnapshot,
    EraEnum,
    AgentSchema,
    ResourceNode,
    SettlementNode,
    HouseNode,
    GravestoneNode,
    SimulationLog,
    LogCategoryEnum,
)

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

SAVE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "saves"))
LEGACY_SAVE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "engine", "saves"))


def _sanitize_save_name(name: str) -> str:
    normalized = re.sub(r"\s+", "_", name.strip())
    normalized = re.sub(r"[^a-zA-Z0-9_-]", "", normalized)
    return normalized[:80] or "save"


def _ensure_unique_filename(base_name: str, ext: str) -> str:
    candidate = f"{base_name}{ext}"
    if not os.path.exists(os.path.join(SAVE_DIR, candidate)):
        return candidate

    counter = 2
    while True:
        candidate = f"{base_name}_{counter}{ext}"
        if not os.path.exists(os.path.join(SAVE_DIR, candidate)):
            return candidate
        counter += 1


def _safe_filename(filename: str) -> str:
    return os.path.basename(filename)


def _build_snapshot_filename(engine, label: str = "", save_name: Optional[str] = None) -> str:
    if save_name:
        safe = _sanitize_save_name(save_name)
        return _ensure_unique_filename(safe, ".json")

    prefix = engine.active_model.replace(".", "_").replace(":", "_")
    base = f"{prefix}_{label}_tick{engine.tick}".strip("_")
    return _ensure_unique_filename(base, ".json")


def load_snapshot(engine, filename: str) -> bool:
    safe_name = _safe_filename(filename)
    fpath = os.path.join(SAVE_DIR, safe_name)
    if not os.path.exists(fpath):
        legacy_path = os.path.join(LEGACY_SAVE_DIR, safe_name)
        if os.path.exists(legacy_path):
            fpath = legacy_path
        else:
            return False

    try:
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)

        schema_version = int(data.get("schema_version", 1))

        engine.tick = data.get("tick", 0)
        engine.era = EraEnum(data.get("era", EraEnum.PREHISTORIC.value))
        model_name = data.get("model_name", "qwen2.5:1.5b")
        engine.active_model = {
            "llama3.1": "qwen2.5:1.5b",
            "llama3.1:latest": "qwen2.5:1.5b",
            "qwen2.5": "qwen2.5:1.5b",
            "qwen3": "qwen2.5:1.5b",
            "qwen3:4b": "qwen3:4b",
            "gemma": "gemma3:4b",
            "gemma3:4b": "gemma3:4b",
            "mistral": "deepseek-coder:1.3b",
            "mistral:latest": "deepseek-coder:1.3b",
            "deepseek-coder:1.3b": "deepseek-coder:1.3b",
        }.get(model_name, model_name)
        engine.violence_level = str(data.get("violence_level", getattr(engine, "violence_level", "normal"))).lower()

        engine.terrain = data.get("terrain", [])
        engine.original_terrain = [row[:] for row in engine.terrain]
        if engine.terrain:
            engine.map_size = len(engine.terrain)

        engine.agents = {aid: AgentSchema(**adata) for aid, adata in data.get("agents", {}).items()}
        for _agent in engine.agents.values():
            if hasattr(engine, "_ensure_agent_skills"):
                engine._ensure_agent_skills(_agent)
            if not getattr(_agent, "royal_title", None):
                _agent.royal_title = ""
        engine.resources = [ResourceNode(**r) for r in data.get("resources", [])]
        engine.settlements = [SettlementNode(**s) for s in data.get("settlements", [])]
        engine.houses = [HouseNode(**h) for h in data.get("houses", [])]
        engine.gravestones = [GravestoneNode(**g) for g in data.get("gravestones", [])]
        engine.relationships = data.get("relationships", {})
        engine.relationship_details = data.get("relationship_details", {})
        engine.groundwater = data.get("groundwater", {})
        engine.weather = data.get("weather", {})
        engine.global_quest = data.get("global_quest")
        engine.quest_history = data.get("quest_history", [])
        engine.next_quest_tick = int(data.get("next_quest_tick", engine.tick + 80))
        if not engine.groundwater and hasattr(engine, "_init_hydrology"):
            engine._init_hydrology()
        engine.logs = [SimulationLog(**l) for l in data.get("logs", [])]

        if schema_version < 2:
            for agent in engine.agents.values():
                # Legacy saves stored hunger as hunger level (0 = full, 100 = starving).
                # Convert it to the new fullness semantics (100 = full, 0 = starving).
                agent.vitals.hunger = float(max(0.0, min(100.0, 100.0 - float(agent.vitals.hunger))))

        if hasattr(engine, "social_pair_last_log"):
            engine.social_pair_last_log.clear()
        if hasattr(engine, "social_links"):
            engine.social_links.clear()
        if hasattr(engine, "log_signature_last_tick"):
            engine.log_signature_last_tick.clear()
        
        engine.pending_cognitive_tasks.clear()
        engine.target_tick = 0

        return True
    except Exception as e:
        print(f"Error loading snapshot: {e}")
        return False


def save_snapshot(
    engine,
    label: str = "",
    overwrite_filename: Optional[str] = None,
    save_name: Optional[str] = None,
) -> str:
    os.makedirs(SAVE_DIR, exist_ok=True)
    snap = SimulationSnapshot(
        schema_version=2,
        model_name=engine.active_model,
        violence_level=str(getattr(engine, "violence_level", "normal")).lower(),
        tick=engine.tick,
        era=engine.era.value,
        agents={aid: a.model_dump(mode="json") for aid, a in engine.agents.items()},
        terrain=engine.terrain,
        resources=[r.model_dump(mode="json") for r in engine.resources],
        settlements=[s.model_dump(mode="json") for s in engine.settlements],
        houses=[h.model_dump(mode="json") for h in engine.houses],
        gravestones=[g.model_dump(mode="json") for g in engine.gravestones],
        relationships=dict(engine.relationships),
        relationship_details=dict(getattr(engine, "relationship_details", {})),
        groundwater=dict(getattr(engine, "groundwater", {})),
        weather=dict(getattr(engine, "weather", {})),
        global_quest=getattr(engine, "global_quest", None),
        quest_history=list(getattr(engine, "quest_history", [])),
        next_quest_tick=int(getattr(engine, "next_quest_tick", 0)),
        logs=[l.model_dump(mode="json") for l in engine.logs],
    )

    if overwrite_filename:
        fname = _safe_filename(overwrite_filename)
        if not fname.endswith(".json"):
            fname = f"{fname}.json"
    else:
        fname = _build_snapshot_filename(engine, label=label, save_name=save_name)

    fpath = os.path.join(SAVE_DIR, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(snap.model_dump(mode="json"), f, indent=2, ensure_ascii=False)

    engine.add_log(LogCategoryEnum.SYSTEM, f"Simulation saved: {fname}")
    return fname


def save_screenshot(base_name: str, data_url: str) -> Optional[str]:
    if not data_url or "," not in data_url:
        return None

    try:
        header, encoded = data_url.split(",", 1)
        if "image/png" not in header:
            return None
        raw = base64.b64decode(encoded)
    except (ValueError, binascii.Error):
        return None

    fname = f"{base_name}.png"
    with open(os.path.join(SAVE_DIR, fname), "wb") as f:
        f.write(raw)
    return fname


def save_excel_report(engine, base_name: str) -> Optional[str]:
    if Workbook is None:
        return None

    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    summary.append(["field", "value"])
    summary.append(["saved_at", datetime.now().isoformat(timespec="seconds")])
    summary.append(["model", engine.active_model])
    summary.append(["tick", engine.tick])
    summary.append(["era", engine.era.value if hasattr(engine.era, "value") else str(engine.era)])
    summary.append(["map_preset", getattr(engine, "map_preset", "unknown")])
    summary.append(["map_size", engine.map_size])
    summary.append(["agent_count_alive", len([a for a in engine.agents.values() if a.is_alive])])
    summary.append(["resource_count", len(engine.resources)])
    summary.append(["settlement_count", len(engine.settlements)])
    summary.append(["house_count", len(engine.houses)])
    if hasattr(engine, "compute_metrics"):
        metrics = engine.compute_metrics()
        summary.append(["gini_coin", metrics.get("gini_coin", 0)])
        summary.append(["social_density", metrics.get("social_density", 0)])
        summary.append(["alliance_edges", metrics.get("alliance_edges", 0)])
        summary.append(["territories", metrics.get("territories", 0)])

    agents_ws = wb.create_sheet("agents")
    agents_ws.append([
        "id",
        "name",
        "gender",
        "x",
        "y",
        "age",
        "max_age",
        "life_phase",
        "social_class",
        "job",
        "is_alive",
        "allies",
        "energy",
        "hunger",
        "hydration",
        "social",
        "wood",
        "food",
        "coin",
        "stone",
        "tools",
        "meat",
        "crop",
        "fruit",
        "herb",
        "has_boat",
        "has_horse",
        "has_cart",
        "has_car",
        "partner_id",
        "house_id",
        "kindness",
        "bravery",
        "sociability",
        "intellect",
        "creativity",
        "ambition",
        "empathy",
        "cunning",
        "skills_json",
        "royal_title",
    ])
    for agent in engine.agents.values():
        agents_ws.append([
            agent.id,
            agent.name,
            getattr(agent.gender, "value", str(agent.gender)),
            agent.x,
            agent.y,
            agent.age,
            agent.max_age,
            getattr(agent.life_phase, "value", str(agent.life_phase)),
            getattr(agent.social_class, "value", str(agent.social_class)),
            agent.job,
            agent.is_alive,
            ",".join(agent.allies),
            agent.vitals.energy,
            agent.vitals.hunger,
            agent.vitals.hydration,
            agent.vitals.social,
            agent.inventory.wood,
            agent.inventory.food,
            agent.inventory.coin,
            agent.inventory.stone,
            agent.inventory.tools,
            agent.inventory.meat,
            agent.inventory.crop,
            agent.inventory.fruit,
            agent.inventory.herb,
            agent.inventory.has_boat,
            agent.inventory.has_horse,
            agent.inventory.has_cart,
            agent.inventory.has_car,
            agent.partner_id,
            agent.house_id,
            agent.personality.kindness,
            agent.personality.bravery,
            agent.personality.sociability,
            agent.personality.intellect,
            agent.personality.creativity,
            agent.personality.ambition,
            agent.personality.empathy,
            agent.personality.cunning,
            json.dumps(getattr(agent, "skills", {}), ensure_ascii=False),
            getattr(agent, "royal_title", ""),
        ])

    resources_ws = wb.create_sheet("resources")
    resources_ws.append(["id", "type", "x", "y"])
    for res in engine.resources:
        resources_ws.append([res.id, getattr(res.type, "value", str(res.type)), res.x, res.y])

    settlements_ws = wb.create_sheet("settlements")
    settlements_ws.append(["id", "owner_id", "x", "y", "is_farming", "territory_radius", "allied_with"])
    for s in engine.settlements:
        settlements_ws.append([
            s.id,
            s.owner_id,
            s.x,
            s.y,
            s.is_farming,
            getattr(s, "territory_radius", 0),
            ",".join(getattr(s, "allied_with", [])),
        ])

    houses_ws = wb.create_sheet("houses")
    houses_ws.append(["id", "owner_id", "type", "x", "y", "level", "residents", "territory_radius", "under_construction"])
    for h in engine.houses:
        houses_ws.append([
            h.id,
            h.owner_id,
            getattr(h.type, "value", str(h.type)),
            h.x,
            h.y,
            h.level,
            ",".join(h.residents),
            h.territory_radius,
            h.is_under_construction,
        ])

    logs_ws = wb.create_sheet("logs")
    logs_ws.append([
        "id",
        "tick",
        "calendar_date",
        "category",
        "interaction_type",
        "participant_ids",
        "relationship_change",
        "source_agent_id",
        "message",
    ])
    for log in engine.logs:
        logs_ws.append([
            log.id,
            log.tick,
            getattr(log, "calendar_date", ""),
            getattr(log.category, "value", str(log.category)),
            getattr(log, "interaction_type", None),
            ",".join(getattr(log, "participant_ids", []) or []),
            getattr(log, "relationship_change", None),
            getattr(log, "source_agent_id", None),
            log.message,
        ])

    fname = f"{base_name}.xlsx"
    wb.save(os.path.join(SAVE_DIR, fname))
    return fname


def delete_save_bundle(filename: str) -> bool:
    safe_name = _safe_filename(filename)
    base, _ = os.path.splitext(safe_name)
    deleted_any = False

    for ext in [".json", ".png", ".xlsx"]:
        for save_dir in [SAVE_DIR, LEGACY_SAVE_DIR]:
            path = os.path.join(save_dir, f"{base}{ext}")
            if os.path.exists(path):
                os.remove(path)
                deleted_any = True

    return deleted_any


def list_saves() -> List[Dict[str, Any]]:
    os.makedirs(SAVE_DIR, exist_ok=True)
    entries: List[Dict[str, Any]] = []
    seen_filenames: set[str] = set()

    for save_dir in [SAVE_DIR, LEGACY_SAVE_DIR]:
        if not os.path.exists(save_dir):
            continue

        json_files = [f for f in os.listdir(save_dir) if f.endswith(".json")]
        for fname in json_files:
            if fname in seen_filenames:
                continue
            seen_filenames.add(fname)

            fpath = os.path.join(save_dir, fname)
            base_name = os.path.splitext(fname)[0]
            model = "-"
            tick = 0
            era = "-"

            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    data = json.load(f)
                model = data.get("model_name", "-")
                tick = data.get("tick", 0)
                era = data.get("era", "-")
            except Exception:
                pass

            entries.append(
                {
                    "filename": fname,
                    "display_name": base_name,
                    "model": model,
                    "tick": tick,
                    "era": era,
                    "updated_at": datetime.fromtimestamp(os.path.getmtime(fpath)).isoformat(timespec="seconds"),
                    "has_image": (
                        os.path.exists(os.path.join(SAVE_DIR, f"{base_name}.png"))
                        or os.path.exists(os.path.join(LEGACY_SAVE_DIR, f"{base_name}.png"))
                    ),
                    "has_excel": (
                        os.path.exists(os.path.join(SAVE_DIR, f"{base_name}.xlsx"))
                        or os.path.exists(os.path.join(LEGACY_SAVE_DIR, f"{base_name}.xlsx"))
                    ),
                }
            )

    entries.sort(key=lambda e: e["updated_at"], reverse=True)
    return entries
